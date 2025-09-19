"""
批量处理2D模式GUI模块
处理批量2D图像处理和分析
"""

import os
import cv2
import numpy as np
import tkinter as tk
import logging
import json
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from utils import logger, natural_sort_key
from gui_base import BaseGUI
from analyzer import batch_analyze_ff_images

class Batch2DGUI(BaseGUI):
    """批量处理2D模式GUI"""
    
    def __init__(self, parent):
        super().__init__(parent)
        
        # 批量处理相关
        self.batch_images_dir = None
        self.batch_labels_dir = None
        self.batch_image_files = []
        
        # 状态变量
        self.status_var = tk.StringVar(value="批量处理2D模式：请先选择图像文件夹")
    
    def on_mouse_click(self, event):
        """禁用鼠标点击绘制功能"""
        pass
    
    def on_mouse_drag(self, event):
        """禁用鼠标拖拽绘制功能"""
        pass
    
    def on_mouse_release(self, event):
        """禁用鼠标释放绘制功能"""
        pass
    
    def delete_last_roi(self):
        """禁用删除ROI功能"""
        pass
    
    def clear_all_roi(self):
        """禁用清空ROI功能"""
        pass
    
    def draw_current_roi(self):
        """禁用绘制当前ROI功能"""
        pass
    
    def display_image(self):
        """显示批量处理图像"""
        if self.ff_image is None:
            return
        
        # 应用旋转
        if self.rotation_angle != 0:
            rotated_image = self.apply_rotation(self.ff_image, self.rotation_angle)
        else:
            rotated_image = self.ff_image.copy()
        
        # 绘制ROI到图像上
        display_image = self.draw_rois_on_image(rotated_image.copy())
        
        # 转换为PIL图像
        if len(display_image.shape) == 3:
            pil_image = Image.fromarray(cv2.cvtColor(display_image, cv2.COLOR_BGR2RGB))
        else:
            pil_image = Image.fromarray(display_image)
        
        # 计算缩放比例
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        
        if canvas_width > 1 and canvas_height > 1:
            scale_x = canvas_width / display_image.shape[1]
            scale_y = canvas_height / display_image.shape[0]
            self.scale_factor = min(scale_x, scale_y, 1.0)
            
            # 计算图像在画布中的位置
            new_width = int(display_image.shape[1] * self.scale_factor)
            new_height = int(display_image.shape[0] * self.scale_factor)
            
            self.image_offset_x = (canvas_width - new_width) // 2
            self.image_offset_y = (canvas_height - new_height) // 2
            
            # 调整图像大小
            pil_image = pil_image.resize((new_width, new_height), Image.Resampling.LANCZOS)
        
        # 转换为Tkinter图像
        self.image_tk = ImageTk.PhotoImage(pil_image)
        
        # 在画布上显示图像
        self.canvas.delete("all")
        self.canvas.create_image(self.image_offset_x, self.image_offset_y, 
                               anchor=tk.NW, image=self.image_tk)
        
        # 更新状态显示
        if self.batch_image_files:
            self.batch_index_var.set(f"{self.batch_current_index + 1}/{len(self.batch_image_files)}")
            current_file = self.batch_image_files[self.batch_current_index]
            if self.batch_labels_dir and self.roi_list:
                self.status_var.set(f"当前图像: {os.path.basename(current_file)} | ROI数量: {len(self.roi_list)}")
            elif self.batch_labels_dir:
                self.status_var.set(f"当前图像: {os.path.basename(current_file)} | 未找到标签文件")
            else:
                self.status_var.set(f"当前图像: {os.path.basename(current_file)} | 请选择标签文件夹以显示标签")
        
        # 更新ROI信息显示（2D批量模式专用）
        self.update_batch_2d_roi_info()
        
    def create_toolbar(self, parent):
        """创建批量处理模式工具栏"""
        # 调用父类方法创建工具栏框架
        super().create_toolbar(parent)
        
        # 文件夹选择
        ttk.Button(self.toolbar, text="选择图像文件夹", command=self.open_batch_images_folder).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(self.toolbar, text="选择标签文件夹", command=self.open_batch_labels_folder).pack(side=tk.LEFT, padx=(0, 5))
        
        
        # 图像导航
        ttk.Button(self.toolbar, text="上一张", command=self.prev_batch_image).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(self.toolbar, text="下一张", command=self.next_batch_image).pack(side=tk.LEFT, padx=(0, 5))
        
        # 处理控制
        self.process_button = ttk.Button(self.toolbar, text="开始处理", command=self.start_batch_processing)
        self.process_button.pack(side=tk.LEFT, padx=(0, 5))
        
        ttk.Button(self.toolbar, text="停止处理", command=self.stop_batch_processing).pack(side=tk.LEFT, padx=(0, 5))
        
        # 结果操作
        ttk.Button(self.toolbar, text="导出结果", command=self.export_batch_results).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(self.toolbar, text="设置输出目录", command=self.set_output_directory).pack(side=tk.LEFT, padx=(0, 5))
        
        # 导出选项
        self.save_individual_json_var = tk.BooleanVar()
        json_check = ttk.Checkbutton(self.toolbar, text="保存单独JSON", variable=self.save_individual_json_var)
        json_check.pack(side=tk.LEFT, padx=(10, 0))
        
        # 调试日志控制
        self.debug_log_var = tk.BooleanVar()
        debug_check = ttk.Checkbutton(self.toolbar, text="调试日志", variable=self.debug_log_var, 
                                    command=self.toggle_debug_log)
        debug_check.pack(side=tk.LEFT, padx=(10, 0))
        
        # 图像信息显示
        self.batch_index_var = tk.StringVar(value="0/0")
        ttk.Label(self.toolbar, textvariable=self.batch_index_var).pack(side=tk.RIGHT, padx=(5, 0))
    
    def toggle_debug_log(self):
        """切换调试日志显示"""
        if hasattr(self, 'debug_log_var'):
            debug_enabled = self.debug_log_var.get()
            if debug_enabled:
                logger.setLevel(logging.DEBUG)
                self.status_var.set("调试日志已启用")
            else:
                logger.setLevel(logging.INFO)
                self.status_var.set("调试日志已禁用")
    
    def handle_mode_specific_keys(self, event):
        """处理批量处理模式特定的快捷键"""
        if event.keysym in ["Left", "a", "A"]:
            self.prev_batch_image()
        elif event.keysym in ["Right", "d", "D"]:
            self.next_batch_image()
        elif event.keysym == "space":
            if self.batch_processing:
                self.stop_batch_processing()
            else:
                self.start_batch_processing()
        elif event.state & 0x4 and event.keysym == "e":  # Ctrl+E
            self.export_batch_results()
    
    def open_batch_images_folder(self):
        """打开批量图像文件夹"""
        folder_path = filedialog.askdirectory(title="选择图像文件夹")
        if folder_path:
            self.batch_images_dir = folder_path
            self.load_batch_images()
    
    def open_batch_labels_folder(self):
        """打开批量标签文件夹"""
        folder_path = filedialog.askdirectory(title="选择标签文件夹（可选）")
        if folder_path:
            self.batch_labels_dir = folder_path
            self.status_var.set(f"标签文件夹已设置为: {folder_path}")
            
            # 如果当前有图像显示，重新加载当前图像以显示标签
            if hasattr(self, 'batch_current_index') and self.batch_image_files:
                self.show_batch_image_by_index(self.batch_current_index)
    
    def load_batch_images(self):
        """加载批量图像列表"""
        if not self.batch_images_dir:
            return
        
        # 获取所有图像文件
        image_extensions = ('.png', '.jpg', '.jpeg', '.tiff', '.tif', '.bmp')
        self.batch_image_files = []
        
        for file in os.listdir(self.batch_images_dir):
            if file.lower().endswith(image_extensions):
                self.batch_image_files.append(file)
        
        # 使用自然排序
        self.batch_image_files.sort(key=natural_sort_key)
        
        if self.batch_image_files:
            self.batch_current_index = 0
            self.show_first_batch_image()
            if self.batch_labels_dir:
                self.status_var.set(f"已加载 {len(self.batch_image_files)} 张图像，标签文件夹已设置")
            else:
                self.status_var.set(f"已加载 {len(self.batch_image_files)} 张图像，请选择标签文件夹以显示标签")
        else:
            messagebox.showwarning("警告", "文件夹中没有找到图像文件")
    
    def show_first_batch_image(self):
        """显示第一张批量图像"""
        if self.batch_image_files:
            self.batch_current_index = 0
            self.show_batch_image_by_index(0)
    
    def start_batch_processing(self):
        """开始批量处理"""
        if not self.batch_image_files:
            messagebox.showwarning("警告", "请先选择图像文件夹")
            return
        
        if not self.output_directory:
            messagebox.showwarning("警告", "请先设置输出目录")
            return
        
        # 检查是否有标签文件夹
        if not self.batch_labels_dir:
            messagebox.showwarning("警告", "请先选择标签文件夹")
            return
        
        self.batch_processing = True
        self.process_button.config(text="停止处理")
        self.status_var.set("开始批量处理...")
        
        # 在新线程中处理
        import threading
        processing_thread = threading.Thread(target=self.process_batch_images)
        processing_thread.daemon = True
        processing_thread.start()
    
    def process_batch_images(self):
        """处理批量图像"""
        try:
            self.batch_results = []
            processed_count = 0
            
            for i, image_file in enumerate(self.batch_image_files):
                if not self.batch_processing:  # 检查是否被停止
                    break
                
                # 更新状态
                self.root.after(0, lambda i=i: self.status_var.set(f"正在处理 {i+1}/{len(self.batch_image_files)}: {image_file}"))
                
                # 加载图像
                image_path = os.path.join(self.batch_images_dir, image_file)
                ff_image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                if ff_image is None:
                    logger.warning(f"无法加载图像: {image_path}")
                    continue
                
                # 查找对应的标签文件
                label_path = self.find_corresponding_label(image_file)
                if not label_path:
                    logger.warning(f"未找到标签文件: {image_file}")
                    continue
                
                # 加载标签并转换为ROI
                roi_mask = cv2.imread(label_path, cv2.IMREAD_GRAYSCALE)
                if roi_mask is None:
                    logger.warning(f"无法加载标签文件: {label_path}")
                    continue
                
                # 将掩码转换为ROI列表
                roi_list = self.convert_mask_to_roi_list(roi_mask)
                if not roi_list:
                    logger.warning(f"标签文件中没有ROI: {label_path}")
                    continue
                
                # 计算每个ROI的脂肪分数
                image_results = []
                for j, roi in enumerate(roi_list):
                    try:
                        # 计算脂肪分数
                        fat_fraction = self.calculate_roi_fat_fraction(ff_image, roi)
                        roi['fat_fraction'] = fat_fraction
                        image_results.append(roi)
                        logger.info(f"ROI {j+1} 脂肪分数: {fat_fraction['fat_fraction']:.3f}")
                    except Exception as e:
                        logger.error(f"计算ROI {j+1} 脂肪分数失败: {str(e)}")
                
                if image_results:
                    result = {
                        'image_file': image_file,
                        'image_path': image_path,
                        'label_path': label_path,
                        'roi_count': len(image_results),
                        'roi_data': image_results,
                        'mean_fat_fraction': np.mean([roi['fat_fraction']['fat_fraction'] for roi in image_results])
                    }
                    self.batch_results.append(result)
                    processed_count += 1
                
                # 更新进度
                self.root.after(0, lambda i=i: self.batch_index_var.set(f"{i+1}/{len(self.batch_image_files)}"))
            
            # 更新UI
            self.root.after(0, self.batch_processing_completed)
            
        except Exception as e:
            logger.error(f"批量处理失败: {str(e)}")
            self.root.after(0, lambda: self.batch_processing_error(str(e)))
    
    def calculate_roi_fat_fraction(self, ff_image, roi):
        """计算ROI的脂肪分数（2D批量模式专用，返回详细统计信息）"""
        try:
            # 根据ROI类型提取像素值
            if roi['type'] == 'rectangle':
                x1, y1 = int(roi['start'][0]), int(roi['start'][1])
                x2, y2 = int(roi['end'][0]), int(roi['end'][1])
                roi_pixels = ff_image[y1:y2, x1:x2]
            elif roi['type'] == 'circle':
                center = (int(roi['center'][0]), int(roi['center'][1]))
                radius = int(roi['radius'])
                # 创建圆形掩码
                mask = np.zeros(ff_image.shape, dtype=np.uint8)
                cv2.circle(mask, center, radius, 1, -1)
                roi_pixels = ff_image[mask == 1]
            elif roi['type'] == 'polygon':
                # 创建多边形掩码
                mask = np.zeros(ff_image.shape, dtype=np.uint8)
                points = np.array(roi['points'], dtype=np.int32)
                cv2.fillPoly(mask, [points], 1)
                roi_pixels = ff_image[mask == 1]
            else:
                return {
                    'fat_fraction': 0.0,
                    'mean_fat_fraction': 0.0,
                    'std_fat_fraction': 0.0,
                    'median_fat_fraction': 0.0,
                    'min_fat_fraction': 0.0,
                    'max_fat_fraction': 0.0,
                    'pixel_count': 0,
                    'coverage_percentage': 0.0,
                    'normalized': False
                }
            
            if len(roi_pixels) == 0:
                return {
                    'fat_fraction': 0.0,
                    'mean_fat_fraction': 0.0,
                    'std_fat_fraction': 0.0,
                    'median_fat_fraction': 0.0,
                    'min_fat_fraction': 0.0,
                    'max_fat_fraction': 0.0,
                    'pixel_count': 0,
                    'coverage_percentage': 0.0,
                    'normalized': False
                }
            
            # 计算脂肪分数
            pixel_values = roi_pixels.flatten()
            pixel_values = pixel_values[pixel_values > 0]  # 排除背景
            
            if len(pixel_values) == 0:
                return {
                    'fat_fraction': 0.0,
                    'mean_fat_fraction': 0.0,
                    'std_fat_fraction': 0.0,
                    'median_fat_fraction': 0.0,
                    'min_fat_fraction': 0.0,
                    'max_fat_fraction': 0.0,
                    'pixel_count': 0,
                    'coverage_percentage': 0.0,
                    'normalized': False
                }
            
            # 检测图像像素值范围并决定是否需要归一化
            min_pixel = np.min(ff_image)
            max_pixel = np.max(ff_image)
            
            # 判断是否需要归一化
            if max_pixel <= 100 and min_pixel >= 0:
                normalized_pixel_values = pixel_values.astype(np.float32)
                normalized = False
            else:
                normalized_pixel_values = pixel_values.astype(np.float32) / max_pixel
                normalized = True
            
            # 计算统计信息
            mean_ff = np.mean(normalized_pixel_values)
            std_ff = np.std(normalized_pixel_values)
            median_ff = np.median(normalized_pixel_values)
            min_ff = np.min(normalized_pixel_values)
            max_ff = np.max(normalized_pixel_values)
            pixel_count = len(normalized_pixel_values)
            
            # 计算覆盖率（ROI面积占图像总面积的比例）
            total_pixels = ff_image.shape[0] * ff_image.shape[1]
            coverage_percentage = (pixel_count / total_pixels) * 100
            
            return {
                'fat_fraction': float(mean_ff),  # 用于显示的主要值
                'mean_fat_fraction': float(mean_ff),
                'std_fat_fraction': float(std_ff),
                'median_fat_fraction': float(median_ff),
                'min_fat_fraction': float(min_ff),
                'max_fat_fraction': float(max_ff),
                'pixel_count': int(pixel_count),
                'coverage_percentage': float(coverage_percentage),
                'normalized': bool(normalized)
            }
            
        except Exception as e:
            logger.error(f"计算脂肪分数失败: {str(e)}")
            return {
                'fat_fraction': 0.0,
                'mean_fat_fraction': 0.0,
                'std_fat_fraction': 0.0,
                'median_fat_fraction': 0.0,
                'min_fat_fraction': 0.0,
                'max_fat_fraction': 0.0,
                'pixel_count': 0,
                'coverage_percentage': 0.0,
                'normalized': False
            }
    
    def batch_processing_completed(self):
        """批量处理完成"""
        self.batch_processing = False
        self.process_button.config(text="开始处理")
        self.status_var.set(f"批量处理完成，共处理 {len(self.batch_results)} 张图像")
        messagebox.showinfo("完成", f"批量处理完成，共处理 {len(self.batch_results)} 张图像")
    
    def batch_processing_error(self, error_msg):
        """批量处理错误"""
        self.batch_processing = False
        self.process_button.config(text="开始处理")
        self.status_var.set(f"批量处理失败: {error_msg}")
        messagebox.showerror("错误", f"批量处理失败: {error_msg}")
    
    def stop_batch_processing(self):
        """停止批量处理"""
        self.batch_processing = False
        self.process_button.config(text="开始处理")
        self.status_var.set("批量处理已停止")
    
    def prev_batch_image(self):
        """上一张批量图像"""
        if not self.batch_image_files:
            return
        
        self.batch_current_index = (self.batch_current_index - 1) % len(self.batch_image_files)
        self.show_batch_image_by_index(self.batch_current_index)
    
    def next_batch_image(self):
        """下一张批量图像"""
        if not self.batch_image_files:
            return
        
        self.batch_current_index = (self.batch_current_index + 1) % len(self.batch_image_files)
        self.show_batch_image_by_index(self.batch_current_index)
    
    def show_batch_image_by_index(self, index):
        """显示指定索引的批量图像"""
        if not self.batch_image_files or index >= len(self.batch_image_files):
            return
        
        try:
            # 加载图像
            image_file = self.batch_image_files[index]
            image_path = os.path.join(self.batch_images_dir, image_file)
            
            self.ff_image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
            if self.ff_image is None:
                raise ValueError(f"无法加载图像: {image_path}")
            
            # 初始化ROI列表
            self.roi_list = []
            self.roi_mask = None
            
            # 如果有标签文件夹，则查找对应的标签文件
            if self.batch_labels_dir:
                logger.debug(f"开始查找标签文件: {image_file}")
                logger.debug(f"batch_labels_dir: {self.batch_labels_dir}")
                label_path = self.find_corresponding_label(image_file)
                logger.debug(f"find_corresponding_label返回: {label_path}")
                if label_path:
                    self.roi_mask = cv2.imread(label_path, cv2.IMREAD_GRAYSCALE)
                    if self.roi_mask is not None:
                        # 将掩码转换为ROI列表
                        self.roi_list = self.convert_mask_to_roi_list(self.roi_mask)
                        
                        # 计算每个ROI的脂肪分数
                        for roi in self.roi_list:
                            fat_fraction = self.calculate_roi_fat_fraction(self.ff_image, roi)
                            roi['fat_fraction'] = fat_fraction
                        
                        logger.info(f"从标签文件加载了 {len(self.roi_list)} 个ROI: {os.path.basename(label_path)}")
                    else:
                        logger.warning(f"无法加载标签文件: {label_path}")
                else:
                    logger.info(f"未找到对应的标签文件: {image_file}")
            else:
                logger.info("未选择标签文件夹，仅显示图像")
            
            # 重置旋转角度
            self.rotation_angle = 0
            
            # 显示图像
            self.display_image()
            self.update_batch_index()
            
            # 显示处理结果（如果有）
            if hasattr(self, 'batch_results') and index < len(self.batch_results):
                self.display_batch_result(index)
            
            self.status_var.set(f"显示图像: {image_file}")
            
        except Exception as e:
            messagebox.showerror("错误", f"加载图像失败: {str(e)}")
            logger.error(f"加载图像失败: {str(e)}")
    
    def display_batch_result(self, index):
        """显示批量处理结果"""
        if index >= len(self.batch_results):
            return
        
        result = self.batch_results[index]
        
        # 显示结果信息
        if 'mean_fat_fraction' in result:
            self.status_var.set(f"平均脂肪分数: {result['mean_fat_fraction']:.3f}")
        
        # 更新ROI信息
        if 'roi_area' in result:
            self.roi_info_var.set(f"ROI面积: {result['roi_area']} 像素")
    
    def update_batch_index(self):
        """更新批量处理索引显示"""
        if self.batch_image_files:
            self.batch_index_var.set(f"{self.batch_current_index + 1}/{len(self.batch_image_files)}")
        else:
            self.batch_index_var.set("0/0")
    
    def update_batch_2d_roi_info(self):
        """更新2D批量模式ROI信息显示（独立于其他模式）"""
        if hasattr(self, 'output_text') and self.output_text:
            self.output_text.delete(1.0, tk.END)
            
            if not self.roi_list:
                self.output_text.insert(tk.END, "暂无ROI数据\n\n请选择标签文件夹以显示ROI")
                return
            
            # 显示所有ROI的计算结果
            output_lines = []
            output_lines.append(f"ROI总数: {len(self.roi_list)}\n")
            output_lines.append("=" * 40 + "\n")
            
            for i, roi in enumerate(self.roi_list):
                output_lines.append(f"ROI #{i+1} ({roi['type']}):\n")
                
                if 'fat_fraction' in roi and roi['fat_fraction'] is not None:
                    ff = roi['fat_fraction']
                    
                    # 2D批量模式：fat_fraction现在是详细统计字典
                    if isinstance(ff, dict):
                        output_lines.append(f"  脂肪分数: {ff['fat_fraction']:.3f}\n")
                        
                        # 分布信息
                        output_lines.append("\n【分布信息】\n")
                        range_val = ff['max_fat_fraction'] - ff['min_fat_fraction']
                        output_lines.append(f"  数值范围: {range_val:.3f}\n")
                        mean_val = ff['mean_fat_fraction']
                        std_val = ff['std_fat_fraction']
                        cv = (std_val / mean_val) * 100 if mean_val > 0 else 0
                        output_lines.append(f"  变异系数: {cv:.2f}%\n")
                        
                        # 区域信息
                        output_lines.append("\n【区域信息】\n")
                        output_lines.append(f"  像素数量: {ff['pixel_count']:,}\n")
                        output_lines.append(f"  覆盖率: {ff['coverage_percentage']:.2f}%\n")
                        
                        # 质量评估
                        output_lines.append("\n【质量评估】\n")
                        std_val = ff['std_fat_fraction']
                        if std_val < 0.05:
                            quality = "优秀"
                        elif std_val < 0.1:
                            quality = "良好"
                        elif std_val < 0.2:
                            quality = "一般"
                        else:
                            quality = "较差"
                        output_lines.append(f"  数据质量: {quality}\n")
                        output_lines.append(f"  归一化: {'是' if ff['normalized'] else '否'}\n")
                    else:
                        # 兼容旧格式（简单数值）
                        output_lines.append(f"  脂肪分数: {ff:.3f}\n")
                        output_lines.append("\n【分布信息】\n")
                        output_lines.append(f"  数值范围: 无法计算\n")
                        output_lines.append(f"  变异系数: 无法计算\n")
                        output_lines.append("\n【区域信息】\n")
                        output_lines.append(f"  像素数量: 无法计算\n")
                        output_lines.append(f"  覆盖率: 无法计算\n")
                        output_lines.append("\n【质量评估】\n")
                        output_lines.append(f"  数据质量: 无法评估\n")
                        output_lines.append(f"  归一化: 否\n")
                else:
                    output_lines.append("  未计算脂肪分数\n")
                
                output_lines.append("=" * 40 + "\n")
            
            self.output_text.insert(tk.END, "".join(output_lines))
    
    def export_batch_results(self):
        """导出批量处理结果"""
        if not self.batch_results:
            messagebox.showwarning("警告", "没有结果可导出")
            return
        
        if not self.output_directory:
            messagebox.showwarning("警告", "请先设置输出目录")
            return
        
        try:
            # 保存每幅图的JSON文件（可选）
            if self.save_individual_json_var.get():
                for result in self.batch_results:
                    base_name = os.path.splitext(result['image_file'])[0]
                    json_path = os.path.join(self.output_directory, f"{base_name}_roi_data.json")
                    
                    # 准备JSON数据
                    roi_data = []
                    for roi in result['roi_data']:
                        if isinstance(roi['fat_fraction'], dict):
                            fat_fraction_value = roi['fat_fraction']['fat_fraction']
                        else:
                            fat_fraction_value = roi['fat_fraction'] if isinstance(roi['fat_fraction'], (int, float)) else 0
                        
                        roi_data.append({
                            'type': roi['type'],
                            'fat_fraction': fat_fraction_value,
                            'start': roi.get('start', []),
                            'end': roi.get('end', []),
                            'center': roi.get('center', []),
                            'radius': roi.get('radius', 0),
                            'points': roi.get('points', [])
                        })
                    
                    json_data = {
                        'image_file': result['image_file'],
                        'roi_count': result['roi_count'],
                        'roi_data': roi_data,
                        'mean_fat_fraction': result['mean_fat_fraction']
                    }
                    
                    with open(json_path, 'w', encoding='utf-8') as f:
                        json.dump(json_data, f, indent=2, ensure_ascii=False)
                    
                    logger.info(f"已保存JSON文件: {json_path}")
            
            # 创建汇总表格
            import pandas as pd
            
            summary_data = []
            for result in self.batch_results:
                for i, roi in enumerate(result['roi_data']):
                    if isinstance(roi['fat_fraction'], dict):
                        fat_fraction_value = roi['fat_fraction']['fat_fraction']
                    else:
                        fat_fraction_value = roi['fat_fraction'] if isinstance(roi['fat_fraction'], (int, float)) else 0
                    
                    summary_data.append({
                        '图像文件': result['image_file'],
                        'ROI编号': i + 1,
                        'ROI类型': roi['type'],
                        '脂肪分数': fat_fraction_value,
                        '平均脂肪分数': result['mean_fat_fraction'],
                        'ROI总数': result['roi_count']
                    })
            
            df = pd.DataFrame(summary_data)
            
            # 优化Excel表格显示，合并重复行
            excel_path = os.path.join(self.output_directory, "batch_analysis_results.xlsx")
            try:
                with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='ROI分析结果', index=False)
                    
                    # 获取工作表对象
                    worksheet = writer.sheets['ROI分析结果']
                    
                    # 合并相同图像文件的单元格
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    from openpyxl.styles import Alignment
                    
                    # 重新写入数据并合并单元格
                    worksheet.delete_rows(1, worksheet.max_row)
                    
                    # 写入标题行
                    headers = ['图像文件', 'ROI编号', 'ROI类型', '脂肪分数', '平均脂肪分数', 'ROI总数']
                    for col, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.value = header
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # 写入数据并合并相同图像文件的单元格
                    current_row = 2
                    current_image = None
                    image_start_row = 2
                    
                    for _, row in df.iterrows():
                        if current_image != row['图像文件']:
                            # 如果之前有图像需要合并，先合并
                            if current_image is not None and current_row > image_start_row:
                                worksheet.merge_cells(f'A{image_start_row}:A{current_row-1}')
                                worksheet.merge_cells(f'E{image_start_row}:E{current_row-1}')
                                worksheet.merge_cells(f'F{image_start_row}:F{current_row-1}')
                            
                            # 开始新的图像
                            current_image = row['图像文件']
                            image_start_row = current_row
                        
                        # 写入当前行数据
                        for col, value in enumerate(row, 1):
                            cell = worksheet.cell(row=current_row, column=col)
                            cell.value = value
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        current_row += 1
                    
                    # 合并最后一个图像的单元格
                    if current_image is not None and current_row > image_start_row:
                        worksheet.merge_cells(f'A{image_start_row}:A{current_row-1}')
                        worksheet.merge_cells(f'E{image_start_row}:E{current_row-1}')
                        worksheet.merge_cells(f'F{image_start_row}:F{current_row-1}')
                    
                    # 设置列宽
                    column_widths = [25, 10, 12, 12, 15, 10]
                    for col, width in enumerate(column_widths, 1):
                        worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width
                
                logger.info(f"已保存Excel文件: {excel_path}")
            except Exception as e:
                # 如果Excel保存失败，保存CSV
                csv_path = os.path.join(self.output_directory, "batch_analysis_results.csv")
                df.to_csv(csv_path, index=False, encoding='utf-8-sig')
                logger.info(f"已保存CSV文件: {csv_path}")
                logger.warning(f"Excel保存失败，已保存CSV: {str(e)}")
            
            # 不再保存总的JSON结果文件
            
            self.status_var.set(f"结果已导出到: {self.output_directory}")
            messagebox.showinfo("成功", f"结果已导出到: {self.output_directory}")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败: {str(e)}")
            logger.error(f"导出失败: {str(e)}")
